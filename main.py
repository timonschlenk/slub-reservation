#import Selenium Libaries
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Import the other libraries
import time
import schedule
import json
import os
from openpyxl import load_workbook  # Import openpyxl to read Excel files
from datetime import datetime

# Load data from Excel file
excel_path = os.path.join(os.path.dirname(__file__), 'data.xlsx')
workbook = load_workbook(excel_path)
sheet = workbook.active

def generate_link(date:str, start_time:str, end_time:str, room_number="0.66"):
    # Generate a link for the reservation page with the given date and time
    formated_start_time = start_time.replace(':', '%3A')
    formated_end_time = end_time.replace(':', '%3A')
    base_url = "https://raumbuchung.slub-dresden.de/Web/reservation/"
    formated_room_number = 18 if room_number == "0.40" else 19 if room_number == "0.42" else 20 if room_number == "0.43" else 21 if room_number == "0.46" else 22 if room_number == "0.47" else 23
    level_number = 1
    return f"{base_url}?rid={formated_room_number}&sid={level_number}&rd={date}&sd={date}{formated_start_time}&ed={date}{formated_end_time}"

def make_reservation(date:str, start_time:str, end_time:str, room_number=20, title="Physik Lerngruppe", description="Hausaufgaben", person_count=1, pause=1):
    LINK = generate_link(date=date, start_time=start_time, end_time=end_time, room_number=room_number)  # Generate the link for the reservation page

    # Setup Chrome options and service
    service = Service(executable_path='./chromedriver.exe')
    driver = webdriver.Chrome(service=service)

    # Load credentials from config file using a relative path
    config_path = os.path.join(os.path.dirname(__file__), './config.json')
    with open(config_path, 'r') as config_file:
        config = json.load(config_file)

    USERNAME = config['username']
    PASSWORD = config['password']

    # Execute the script
    driver.get(LINK)  # Open the reservation page
    SubmitButton = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CLASS_NAME, 'btn-large'))
    )  # Wait for the submit button to be clickable
    SubmitButton.click()  # Click the submit button

    UsernameInput = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'txtUsername'))
    )  # Wait for the username input field to be present

    PasswordInput = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'txtPassword'))
    )  # Wait for the password input field to be present 

    SubmitButton = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'btnLogin'))
    )  # Wait for the submit button to be present

    UsernameInput.send_keys(USERNAME)  # Enter the username
    PasswordInput.send_keys(PASSWORD)  # Enter the password
    SubmitButton.click()  # Click the submit button

    TitleInput = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'reservation-title'))
    )  # Wait for the title input field to be present

    DescriptionInput = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'reservation-description'))
    )  # Wait for the description input field to be present

    PersonCountInput = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'react-select-2-input'))
    )  # Wait for the person count input field to be present

    ReservationTermsCheckbox = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'reservation-terms-checkbox'))
    )  # Wait for the privacy checkbox to be present

    SubmitButton2 = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CLASS_NAME, 'btn-primary'))
    )  # Wait for the submit button to be present

    TitleInput.send_keys(title)  # Enter the title
    DescriptionInput.send_keys(description)  # Enter the description
    PersonCountInput.send_keys(f"{person_count}" + Keys.ENTER)  # Enter the person count
    ReservationTermsCheckbox.click()  # Click the privacy checkbox
    SubmitButton2.click()  # Click the submit button

    # Wait for ... seconds
    time.sleep(pause)
    sucess = True if driver.find_elements(By.CLASS_NAME, 'success') else False  # Check if the reservation was successful
    # Close the browser
    driver.quit()
    return sucess  # Return the success status

def run_reservation_script():
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=8, values_only=True), start=2):
        DATE = row[0]                           # Column A: Date
        START_TIME = row[1]                     # Column B: Start Time
        END_TIME = row[2]                       # Column C: End Time
        ROOM_NUMBERS = row[3].split("|")        # Column D: Room Number                  # Column E: Level Number
        TITLE = row[4]                          # Column E: Title
        DESCRIPTION = row[5]                    # Column F: Description
        PERSON_COUNT = row[6]                   # Column G: Person Count
        STATUS = row[7]                         # Column H: Status
        current_date = datetime.now().date()    # Get the current date
        formated_date = datetime.strptime(DATE, '%Y-%m-%d').date()
        
        if STATUS in ["Success - Reservation for Room 0.40", "Success - Reservation for Room 0.42", "Success - Reservation for Room 0.43", "Success - Reservation for Room 0.46", "Success - Reservation for Room 0.47","Success - Reservation for Room 0.66", "Failed for all Rooms", "Skipped (Past)"]:
            print(f"Skipping row {row_index} with status: {STATUS}")  # Skip rows with these statuses
            continue
        if (formated_date - current_date).days < 0:
            sheet[f"H{row_index}"] = "Skipped (Past)"  # Write status in column H
            print(f"Skipping row {row_index} with status: Skipped (Past)")  # Skip past dates
            continue
        if (formated_date - current_date).days > 14:
            sheet[f"H{row_index}"] = "Skipped (Future)"  # Write status in column H
            print(f"Skipping row {row_index} with status: Skipped (Future)")
            continue

        for ROOM_NUMBER in ROOM_NUMBERS:
            success = make_reservation(date=DATE, start_time=START_TIME, end_time=END_TIME, room_number=ROOM_NUMBER, title=TITLE, description=DESCRIPTION, person_count=PERSON_COUNT)  # Call the function to make the reservation
            if success:
                sheet[f"H{row_index}"] = f"Success - Reservation for Room {ROOM_NUMBER}"  # Write success status in column I
                print(f"Successfully made reservation for row {row_index} for Room {ROOM_NUMBER}")
                break  # Break the loop if the reservation was successful
        if not success:
            sheet[f"H{row_index}"] = "Failed for all Rooms"  # Write failure status in column H
            print(f"Failed to make reservation for row {row_index} for all rooms")
    workbook.save(excel_path)  # Save the workbook after making reservations
    workbook.close()  # Close the workbook
    print("Reservations completed. Workbook saved.")  # Print completion message

run_reservation_script()  # Run the script once at startup

# Schedule the script to run at every full hour
schedule.every().hour.at(":00").do(run_reservation_script)

print("Scheduler is running. Press Ctrl+C to stop.")

# Keep the script running
while True:
    schedule.run_pending()
    time.sleep(1)